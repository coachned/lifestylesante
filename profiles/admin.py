from django.contrib import admin
from .models import Employe, MarquerArrivee, MarquerDepart, RapportMensuel
import csv
from django.http import HttpResponse
from django.utils.text import slugify
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from io import BytesIO
from datetime import datetime, date
from django.conf import settings

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


# Export Excel
def export_as_excel(modeladmin, request, queryset):
    meta = modeladmin.model._meta
    field_names = [field.name for field in meta.fields if field.name not in ['is_deleted', 'id']]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = slugify(str(meta.verbose_name_plural))  # <-- Correction ici

    # En-têtes en gras
    for col_num, field_name in enumerate(field_names, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = field_name
        cell.font = Font(bold=True)

    # Données
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field_name in enumerate(field_names, 1):
            value = getattr(obj, field_name)
        # Formatage de tous les champs de type date/datetime
        if isinstance(value, (datetime, date)):
            value = value.strftime('%Y-%m-%d %H:%M')
        worksheet.cell(row=row_num, column=col_num, value=str(value))

    # Largeur colonnes
    for col_num, field_name in enumerate(field_names, 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = max(20, len(field_name) + 5)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"{slugify(str(meta.verbose_name_plural))}.xlsx"  # <-- Correction ici
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

export_as_excel.short_description = "Exporter en Excel"


# Export CSV
def export_as_csv(modeladmin, request, queryset):
    meta = modeladmin.model._meta
    field_names = [field.name for field in meta.fields if field.name not in ['is_deleted', 'id']]

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename={slugify(str(meta.verbose_name_plural))}.csv'  # <-- Correction ici

    writer = csv.writer(response)
    writer.writerow(field_names)

    for obj in queryset:
        row = []
    for field in field_names:
        value = getattr(obj, field)
        # Si c'est un datetime ou une date, on le formate proprement
        if isinstance(value, (datetime, date)):
            value = value.strftime('%Y-%m-%d %H:%M')
        row.append(str(value))
    writer.writerow(row)

    return response

export_as_csv.short_description = "Exporter en CSV"


# Export PDF
def export_as_pdf(modeladmin, request, queryset):
    meta = modeladmin.model._meta
    field_names = [field.name for field in meta.fields if field.name not in ['is_deleted', 'id']]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    logo_path = settings.BASE_DIR / 'static' / 'img' / 'logo_lifestyle.jpg'
    logo = Image(logo_path, width=100, height=50)
    elements.append(logo)

    styles = getSampleStyleSheet()
    title = Paragraph(f"<b>Export de {str(meta.verbose_name_plural).title()}</b>", styles['Title'])  # <-- Correction ici
    date_gen = Paragraph(f"Date de génération : {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
    elements.extend([title, date_gen, Spacer(1, 12)])

    data = [field_names]
    total_montant = 0

    for obj in queryset:
        row = []
    for field in field_names:
        value = getattr(obj, field)

        # Gérer tous les champs de type date ou datetime
        if isinstance(value, (datetime, date)):
            value = value.strftime('%Y-%m-%d %H:%M')

        # Total montant
        if field == "montant":
            try:
                total_montant += float(value)
            except (ValueError, TypeError):
                pass

        row.append(str(value))
    data.append(row)

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"<b>Total général des montants :</b> {total_montant:.2f} FCFA", styles['Normal']))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Exporté automatiquement depuis l'administration Lifestyle.", styles['Italic']))

    doc.build(elements)
    buffer.seek(0)

    filename = f"{slugify(str(meta.verbose_name_plural))}.pdf"  # <-- Correction ici
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

export_as_pdf.short_description = "Exporter en PDF"


# Admin personnalisée
class BaseAdmin(admin.ModelAdmin):
    actions = [export_as_csv, export_as_excel, export_as_pdf]
    list_filter = ('is_deleted',)

#  Admin Employé
@admin.register(Employe)
class EmployeAdmin(BaseAdmin):
    list_display = ('user', 'nom', 'prenom', 'sexe', 'fonction', 'telephone', 'email')
    search_fields = ('nom', 'prenom', 'telephone', 'email', 'fonction')
    list_filter = ('sexe', 'fonction', 'departement')

#  Admin MarquerArrivee
@admin.register(MarquerArrivee)
class MarquerArriveeAdmin(BaseAdmin):
    list_display = ('employe', 'date_arrivee', 'arrivee', 'montant')
    list_filter = ('arrivee', 'date_arrivee')
    search_fields = ('employe__nom', 'employe__prenom', 'date_arrivee')
    ordering = ('-date_arrivee',)

#  Admin MarquerDepart
@admin.register(MarquerDepart)
class MarquerDepartAdmin(BaseAdmin):
    list_display = ('employe', 'date_depart', 'depart')
    list_filter = ('depart', 'date_depart')
    search_fields = ('employe__nom', 'employe__prenom', 'date_depart')
    ordering = ('-date_depart',)

#  Admin RapportMensuel
@admin.register(RapportMensuel)
class RapportMensuelAdmin(BaseAdmin):
    list_display = ('employe', 'mois', 'annee', 'total_arrivees', 'total_departs', 'total_montant')
    list_filter = ('mois', 'annee')
    search_fields = ('employe__nom', 'employe__prenom')
    ordering = ('-annee', '-mois')

